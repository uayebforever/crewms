import openpyxl
from openpyxl.workbook import workbook
from openpyxl.worksheet import worksheet
from openpyxl.worksheet.cell_range import CellRange

from typing import Tuple, List


class SkillGridLoaderBase:
    def __init__(self, workbook_filename):
        # type: (str) -> None
        self.workbook = openpyxl.load_workbook(workbook_filename, data_only=True)  # type: workbook.Workbook

        assert "Skills Grid" in self.workbook.sheetnames
        self.skills_grid_sheet = self.workbook["Skills Grid"]  # type: worksheet.Worksheet

        assert "WnS Bill" in self.workbook.sheetnames
        # self.wns_bill = self.workbook["WnS Bill"]  # type: worksheet.Worksheet

        self.defined_names = self.workbook.defined_names
        self.bounding_box = worksheet.CellRange(self.skills_grid_sheet.print_area[0])

    def _get_cells_for_reference(self, reference_name):

        cells = []

        for sheet_name, cell_reference in self.workbook.defined_names[reference_name].destinations:
            ws = self.workbook[sheet_name]
            cells.append(ws[cell_reference])

        return cells

    def iter_cols(self, defined_name: str):
        """Iterate over the named range by columns."""
        sheet_title, range = self.get_range_for_name(defined_name)
        ws = self.workbook[sheet_title]  # type: worksheet.Worksheet
        return ws.iter_cols(min_col=range.min_col, max_col=range.max_col, min_row=range.min_row, max_row=range.max_row)

    def iter_rows(self, defined_name: str):
        """Iterate over the named range by columns."""
        sheet_title, range = self.get_range_for_name(defined_name)
        ws = self.workbook[sheet_title]  # type: worksheet.Worksheet
        return ws.iter_rows(min_col=range.min_col, max_col=range.max_col, min_row=range.min_row, max_row=range.max_row)

    def get_range_for_name(self, defined_name) -> Tuple[str, CellRange]:
        # see https://openpyxl.readthedocs.io/en/stable/defined_names.html
        dests = list(self.defined_names[defined_name].destinations)
        if len(dests) != 1:
            raise Exception("Destination is either multiple ranges or has no range.")
        sheet_title, range_str = dests[0]
        range = CellRange(range_string=range_str)
        return sheet_title, range

    def validate_spreadsheet(self):

        assert "Duties" in self.defined_names
