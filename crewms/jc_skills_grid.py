from collections import defaultdict

# import pandas as pd
# import sqlalchemy

from openpyxl.worksheet import worksheet
from openpyxl.utils import cell

from typing import List, Dict, Any

from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker

from .skill_grid_loader_base import SkillGridLoaderBase

# Set up logging
from . import slogging

log = slogging.getLogger(__name__)
log.setLevel(slogging.DEBUG)


engine = create_engine("sqlite:///:memory:")
Session = sessionmaker(bind=engine)
session = Session()

from .data import *


Base.metadata.create_all(engine)

def append_or_last(list, value):
    # type: (list, Any) -> None
    if value is None:
        list.append(list[-1])
    else:
        list.append(value)

class LastNone:

    def __init__(self):
        self._store = dict()

    def last_if_none(self, id, value):
        if value is None:
            return self._store[id]
        else:
            self._store[id] = value
            return value


class LastOrNone(object):
    def __init__(self):
        self.last_value = None
    def get(self, value):
        if value is None:
            return self.last_value
        else:
            self.last_value = value
            return value

def _get_or_create_evolution(evolution_name):
    try:
        return session.query(Evolution).filter(Evolution.name == evolution_name).one()
    except:
        evolution = Evolution(name=evolution_name)
        session.add(evolution)
        return evolution



class SkillsGridLoader(SkillGridLoaderBase):

    def _load_skills(self):
        # type: () -> Dict[int, Skill]
        # Skills
        # TODO: Fix this to use "SkillsRef" reference.

        skills = dict()  # type: Dict[int, Skill]
        for row in self.skills_grid_sheet.iter_rows(min_col=1, min_row=13,
                                                    max_col=3, max_row=self.bounding_box.max_row):
            category, skill, level = row
            skill = Skill(id=category.row, category=category.value, name=skill.value, level=level.value)

            skills[skill.id] = skill

            session.add(skill)

        session.commit()
        return skills

    def _load_tasks(self):
        # type: () -> Dict[int, Task]
        # Tasks
        tasks = dict()  # type: Dict[int, Task]

        category_last = LastOrNone()
        evolution_last = LastOrNone()
        task_last = LastOrNone()

        for column in self.iter_cols("Duties"):
            cat, evol, dut, name, match_re = column

            if name.value is None:
                continue

            task = Task(id=cat.column,
                        category=str(category_last.get(cat.value)),
                        evolution=str(evolution_last.get(evol.value)),
                        name=str(task_last.get(name.value)),
                        duty_match_re=match_re.value,
                        rank=0) # TODO: Should do a sensible rank.
            tasks[task.id] = task
            log.debug(task)
            session.add(task)
        session.commit()
        return tasks

    def load(self):
        # type: () -> SkillsGrid
        skills = self._load_skills()
        tasks = self._load_tasks()

        self._link_skills_to_tasks()

        skills_grid = SkillsGrid(skills=skills, tasks=tasks)

        return skills_grid

    def _link_skills_to_tasks(self):
        for row in self.skills_grid_sheet.iter_rows(min_col=5, min_row=14,
                                                    max_col=self.bounding_box.max_col,
                                                    max_row=self.bounding_box.max_row):
            for cell in row:
                if cell.value == "Y":
                    # Create relationship
                    skill = skill_by_id(cell.row)
                    task = task_by_id(cell.column)
                    skill.tasks.append(task)
        session.commit()
        return task_by_id

    def get_evolution(self, name: str, category: str) -> Evolution:
        result = session.query(Evolution).filter(Evolution.category == category, Evolution.name == name).one_or_none()
        if result is None:
            result = Evolution(name=name, category=category)
        return result

    def get_duty(self, name: str, evolution: Evolution) -> Duty:
        result = session.query(Duty).filter(Duty.evolution == evolution, Duty.name == name).one_or_none()
        if result is None:
            result = Duty(name=name, evolution=evolution)
        pass


class SkillsGrid:

    def __init__(self, skills, tasks):
        # type: (Dict[int, Skill], Dict[int, Task]) -> None

        self.skills = skills  # type: Dict[int, Skill]
        self.tasks = tasks  # type: Dict[int, Task]

        self.bills = list()




    @property
    def evolutions(self):
        return set(t.evolution for t in session.query(Task).all())

    @property
    def task_categories(self):
        return [row[0] for row in session.execute(
            select([Task.__table__.c.category]).order_by(Task.__table__.c.id).distinct())]

    @property
    def watchcards(self):
        return session.query(WatchCard).all()


    def watchcards_for_bill(self, bill):
        # type: (str) -> List[WatchCard]
        return session.query(WatchCard).filter(WatchCard.bill == bill).order_by(WatchCard.card_number).all()

    def get_watch_and_station_bill_duties(self, wsb_locations):

        # Iterate over watch bills
        for wsb_name in self.bills:
            # start_col = wns_bill_sheet_locations[wsb_name]
            sheet = self.workbook["WnS Bill " + wsb_name]
            sheet_bounds = worksheet.CellRange(sheet.calculate_dimension())
            # Collect evolution names
            evolutions = dict()  # type: Dict[int, str]
            for column in sheet.iter_cols(min_col=6, max_col=sheet_bounds.max_col,
                                          min_row=4, max_row=4):
                cell = column[0]
                if cell.value != "!":
                    evolutions[cell.column] = cell.value
                else:
                    break

            current_crew_category = None

            # print(evolutions)
            # Iterate over watch cards
            watch_cards_seen = set()
            for row in sheet.iter_rows(min_row=5, max_row=sheet_bounds.max_row,
                                       min_col=1, max_col=1):
                cell = row[0]
                if cell.value.startswith("Area: "):
                    current_crew_category = cell.value[len("Area: "):]
                    continue
                if cell.value is not None and str(cell.value).strip() != "":
                    watch_card = watchcard_by_number_and_bill(cell.value, wsb_name)  # type: WatchCard
                    if watch_card is None:
                        print("Watch card %s on Watch and station bill %s not in skills grid." % (cell.value, wsb_name))
                        continue
                    watch_cards_seen.add(watch_card)
                    watch_card.manning_requirements = sheet.cell(row=cell.row, column=2).value
                    watch_card.crew_category = current_crew_category
                    for column in sheet.iter_cols(min_col=6, max_col=6 + len(evolutions) - 1,
                                                  min_row=cell.row, max_row=cell.row):

                        duty_cell = column[0]  # type: worksheet.Cell

                        evolution = _get_or_create_evolution(evolutions[duty_cell.column])

                        watch_card.duties[evolutions[duty_cell.column]] = \
                            Duty(name=duty_cell.value, evolution=evolution)
                    watch_card.name = sheet.cell(cell.row, 3).value

            if set(self.watchcards_for_bill(wsb_name)) != watch_cards_seen:
                print("Watch cards missing from bill %s:" % wsb_name)
                for card in set(self.watchcards_for_bill(wsb_name)).difference(watch_cards_seen):
                    print("    {0.card_number}".format(card))
                    session.delete(card)

        session.commit()


    def skills_for_task_by_id(self, id):
        # type: (int) -> List[Skill]
        # assert id in self.tasks.index, "Unknown task"
        # return self.skills[self.skills[id] == "Y"]

        # task = session.query(Task).filter(Task.id == id).one()

        return self.tasks[id].skills



    def report_watch_bill_tasks(self, watchbill):

        watchcards = session.query(WatchCard).filter(WatchCard.bill == watchbill).order_by(WatchCard.card_number).all()

        report = []

        if len(watchcards) == 0:
            report.append("No watch cards assigned to the '%s' watch bill." % watchbill)

        for card in watchcards:
            assert isinstance(card, WatchCard)
            report.append("Card {}  (required rank: {})".format(card.card_number, card.required_rank))
            for task in card.tasks:
                report.append("   " + str(task))

        return "\n".join(report)

def watchcard_by_number_and_bill(num, bill):
    # type: (str, str) -> WatchCard
    return session.query(WatchCard).filter(
        WatchCard.card_number == num, WatchCard.bill == bill).one_or_none()

def all_crew_cards_for_category_and_bill(crew_category, bill):
    return session.query(WatchCard).filter(
        WatchCard.bill == bill,
        WatchCard.crew_category == crew_category).all()

def skill_by_id(id):
    # type: (int) -> Skill
    return session.query(Skill).filter(Skill.id == id).one()

def task_by_id(id):
    # type: (int) -> Task
    try:
        return session.query(Task).filter(Task.id == id).one()
    except Exception as e:
        log.error("Task ID %s not found", cell.get_column_letter(id))
        raise e
