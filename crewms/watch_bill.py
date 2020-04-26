from collections import defaultdict, _KT, _VT

from sqlalchemy import create_engine, Column, Integer, String, Table, ForeignKey, select
from sqlalchemy.orm import sessionmaker, relationship
from sqlalchemy.orm.collections import attribute_mapped_collection




import openpyxl
from openpyxl.worksheet import worksheet
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.cell_range import CellRange
import openpyxl

from typing import List, Iterable, Tuple, Dict

from .skill_grid_loader_base import SkillGridLoaderBase
from .data import *
from .pyxl_helpers import get_cells_for_reference, get_range_for_reference

EMERGENCY_EVOLUTION_CATEGORY_NAME = "EMERGENCY PARTIES"
SEA_DUTY_CATEGORY_NAME = "SPECIAL SEA DUTIES"


class KeyedDefaultDict(defaultdict):

    def __missing__(self, key: _KT) -> _VT:
        if self.default_factory is not None:
            self[key] = self.default_factory(key)
            return self[key]


class WatchBillLoader:

    def __init__(self, filename):
        self.filename = filename
        self.worksheet = None  # type: Worksheet
        self.column_iterator = None  # type: WatchBillColumnIterator

    def load_watch_bill_from_xlsx(self):
        # type: () -> WatchBill
        workbook = openpyxl.load_workbook(self.filename, data_only=True)  # , read_only=True)

        assert "WATCH BILL" in workbook.sheetnames
        self.worksheet = workbook["WATCH BILL"]

        named_ranges = workbook.defined_names
        watch_bill_bounds = worksheet.CellRange(self.worksheet.print_area[0])

        duty_names = list()

        column_cell_range = get_range_for_reference(workbook, "WatchBillColumns", bounding_box=watch_bill_bounds)

        self.column_iterator = WatchBillColumnIterator(self.worksheet, column_cell_range)

        # create evolutions
        evolutions = dict()
        evolutions.update(self._create_evolutions(EMERGENCY_EVOLUTION_CATEGORY_NAME))
        evolutions.update(self._create_evolutions(SEA_DUTY_CATEGORY_NAME))

        cards = list()

        # Iterate over rows, skipping the blank/invalid ones
        for row in range(column_cell_range.min_row + 1, watch_bill_bounds.max_row):
            row_data = self.load_row(row)
            if self.row_is_station(row_data):
                card = WatchCard()
                card.id = row
                card.card_number = row_data["Crew No."]
                card.name = row_data["Position"]


                self._add_duties_to_card_for_category(card, evolutions, row, SEA_DUTY_CATEGORY_NAME)
                self._add_duties_to_card_for_category(card, evolutions, row, EMERGENCY_EVOLUTION_CATEGORY_NAME)

                print(card.full_report)
                cards.append(card)

        return WatchBill(cards=cards)

    def _create_evolutions(self, category):
        evolutions = dict()  # type: Dict[str, Evolution]
        for _, evolution_name, _ in self.column_iterator.iterate_category(3, category):
            evolutions[evolution_name] = Evolution(name=evolution_name,
                                                   category=category)
        return evolutions

    def _add_duties_to_card_for_category(self, card: WatchCard, evolutions, row, category):
        for _, evolution, duty_name in self.column_iterator.iterate_category(row, category):
            duty = Duty(evolution=evolutions[evolution], name=duty_name, watch_card_id=card.id)
            evolutions[evolution].duties.append(duty)
            card.duties[evolution] = duty

    def row_is_station(self, row):
        # type: (Dict[str, Any]) -> bool
        return (row["Position"] is not None
                and str(row["Position"]).strip() != ""
                and "Watch" not in str(row["Crew No."])
                and str(row["Position"]) != "Position"
                and "".join([str(v) for v in row.values()]).strip() != "")

    def load_row(self, row):
        return {k: v for i, k, v in self.column_iterator.iterate_row(row)}

class WatchBillLoaderFromSkillsGridWorkbook(SkillGridLoaderBase):

    def _reload_data(self):

        generic_tasks = self.load_watch_and_station_bill_assignments()
        self.get_watch_and_station_bill_duties(None)

        for bill in generic_tasks:
            for crew_category, task in generic_tasks[bill].items():
                for card in all_crew_cards_for_category_and_bill(crew_category, bill):
                    card.tasks.append(task)
        session.commit()


    def load_watch_and_station_bill_assignments(self):
        # Watch and Station Bill Assignments
        wsb_region = dict()
        wsb_region["min row"] = self._get_cells_for_reference("BillAssignmentRef")[0].row
        wsb_region["max row"] = self._get_cells_for_reference("SkillsRef")[0].row - 1
        wsb_region["header col"] = self._get_cells_for_reference("BillAssignmentRef")[0].column
        wsb_locations = dict()
        for row in self.skills_grid_sheet.iter_rows(
            min_row=wsb_region["min row"], min_col=wsb_region["header col"],
            max_row=wsb_region["max row"], max_col=wsb_region["header col"]):
            if row[0].value.startswith("WSB Task Assignment:"):
                wsb_locations[row[0].value[len("WSB Task Assignment:") + 1:]] = row[0].row
        # print(wsb_locations)

        self.bills = [s for s in wsb_locations]

        generic_tasks = defaultdict(dict)  # type: Dict[str, Dict[str, Task]]

        for wsb_name in wsb_locations:
            self.create_watchcards_for_bill(generic_tasks, wsb_locations, wsb_name, wsb_region)

        session.commit()
        return generic_tasks

    def create_watchcards_for_bill(self, generic_tasks, wsb_locations, wsb_name, wsb_region) -> Dict[int, WatchCard]:

        cards = dict()
        for column in self.skills_grid_sheet.iter_cols(min_col=wsb_region["header col"] + 1,
                                                       min_row=wsb_locations[wsb_name],
                                                       max_col=self.bounding_box.max_col,
                                                       max_row=wsb_locations[wsb_name]):
            cell = column[0]
            if cell.value is not None:
                value = str(cell.value)
                if value.startswith("All "):
                    crew_category = value[len("All "):]
                    generic_tasks[wsb_name][crew_category] = task_by_id(cell.column)
                    continue
                for card_id in cell.value.split(","):
                    if card_id not in cards:
                        watch_card = WatchCard(card_number=card_id)
                        cards[card_id] = watch_card
                        session.add(watch_card)
                    else:
                        watch_card = cards[card_id]

                    # print("Creating card %s" % cell.value)
                    watch_card.tasks.append(task_by_id(cell.column))

    def get_watch_and_station_bill_duties(self, wsb_locations):

        # Iterate over watch bills
        for wsb_name in self.bills:
            # start_col = wns_bill_sheet_locations[wsb_name]
            sheet = self.workbook["WnS Bill " + wsb_name]
            sheet_bounds = worksheet.CellRange(sheet.calculate_dimension())
            # Collect evolution names
            evolutions = dict()  # type: Dict[int, str]
            for column in sheet.iter_cols(min_col=6, max_col=sheet_bounds.max_col,
                                          min_row=4, max_row=4):
                cell = column[0]
                if cell.value != "!":
                    evolutions[cell.column] = cell.value
                else:
                    break

            current_crew_category = None

            # print(evolutions)
            # Iterate over watch cards
            watch_cards_seen = set()
            for row in sheet.iter_rows(min_row=5, max_row=sheet_bounds.max_row,
                                       min_col=1, max_col=1):
                cell = row[0]
                if cell.value.startswith("Area: "):
                    current_crew_category = cell.value[len("Area: "):]
                    continue
                if cell.value is not None and str(cell.value).strip() != "":
                    watch_card = watchcard_by_number_and_bill(cell.value, wsb_name)  # type: WatchCard
                    if watch_card is None:
                        print("Watch card %s on Watch and station bill %s not in skills grid." % (cell.value, wsb_name))
                        continue
                    watch_cards_seen.add(watch_card)
                    watch_card.manning_requirements = sheet.cell(row=cell.row, column=2).value
                    watch_card.crew_category = current_crew_category
                    for column in sheet.iter_cols(min_col=6, max_col=6 + len(evolutions) - 1,
                                                  min_row=cell.row, max_row=cell.row):

                        duty_cell = column[0]  # type: worksheet.Cell

                        evolution = _get_or_create_evolution(evolutions[duty_cell.column])

                        watch_card.duties[evolutions[duty_cell.column]] = \
                            Duty(name=duty_cell.value, evolution=evolution)
                    watch_card.name = sheet.cell(cell.row, 3).value

            if set(self.watchcards_for_bill(wsb_name)) != watch_cards_seen:
                print("Watch cards missing from bill %s:" % wsb_name)
                for card in set(self.watchcards_for_bill(wsb_name)).difference(watch_cards_seen):
                    print("    {0.card_number}".format(card))
                    session.delete(card)

        session.commit()



class WatchBill(object):

    def __init__(self, cards):
        # type: (List[WatchCard]) -> None
        self.cards = cards  # type: List[WatchCard]



class WatchBillColumnIterator(object):


    def __init__(self, worksheet, cell_range):
        # type: (Worksheet, CellRange) -> None

        self.worksheet = worksheet  # type: Worksheet
        self.cell_range = cell_range  # type: CellRange
        self.column_index = dict()  # type: Dict[int, Tuple[str, str]]
        self.columns = dict()  # type: Dict[str, Tuple[int, str]]
        self.columns_by_category = defaultdict(list)  # type: Dict[str, List[str]]

        category = ""  # type: str
        for row, col in next(iter(cell_range.rows)):
            category_cell_value = worksheet.cell(row - 1, col).value
            if category_cell_value is not None and str(category_cell_value).strip() != "":
                category = str(category_cell_value)
            if category is None:
                raise Exception("W&S Bill doesn't have a category above the left most column name")
            column_name = worksheet.cell(row, col).value
            self.column_index[col] = (column_name, category_cell_value)
            self.columns[column_name] = (col, category_cell_value)
            self.columns_by_category[category].append(column_name)

    def iterate_row(self, row_to_interate):
        # type: (int) -> Tuple[int, str, str]
        for col in self.worksheet.iter_cols(min_col=self.cell_range.min_col, max_col=self.cell_range.max_col,
                                              min_row=row_to_interate, max_row=row_to_interate):
            cell = col[0]
            assert isinstance(cell, Cell)
            yield cell.column, self.column_index[cell.column][0], cell.value

    # def iterate_emergency_duties(self, row_to_iterate):
    #     # type: (int) -> Tuple[int, str, str]
    #     yield from self.iterate_category(row_to_iterate, EMERGENCY_EVOLUTION_CATEGORY_NAME)
    #
    # def iterate_sea_duties(self, row_to_iterate):
    #     # type: (int) -> Tuple[int, str, str]
    #     yield from self.iterate_category(row_to_iterate, SEA_DUTY_CATEGORY_NAME)

    def iterate_category(self, row_to_iterate, category_name):
        # type: (int, str) -> Tuple[int, str, str]
        for column_name in self.columns_by_category[category_name]:
            column_number = self.columns[column_name][0]
            cell = self.worksheet.cell(row_to_iterate, column_number)
            yield column_number, column_name, cell.value
